


from openpyxl import load_workbook
wb=load_workbook('C:/Users/聂江海/Desktop/python/代码/pythonProject/工作脚本/重庆项目实习生产出台账20201127.xlsx')#加载已有文件
ws=wb.active#启用表
class getAndsend():
   def __init__(self,min_row,max_row, min_col, max_col):
        self.min_row=min_row
        self.max_row=max_row
        self.min_col=min_col
        self.max_col=max_col
        pass

   def getexceldata(self):
      listi=[]
      for i in ws.iter_rows(min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_col, values_only=True):
       listi0=list(i)
       stri=str(listi0)
       stri0=stri
      pass
   pass

   def judgedata(self):
        listi=[]
        for i in ws.iter_rows(min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_col, values_only=True):
         stri=' '
        listi0=list(i) 
        if listi0==['男']:
            stri='先生。' 
            pass
        elif listi0==['女']:
            stri='女士。'
            pass
        else:
            stri='性别异常，请检查'
            pass
   pass

   def glueAndappendstr(self):
        listi=L
        contstantstrf=f
        variablestr=v 
        contstantstrb=b
        comprehensivestr=contstantstrf+variablestr+contstantstrb
        listi.append(comprehensivestr) 
   pass

   def glueAndappenddict(self):
        outsidei=dict(zip(a,b))
        resulti=[]
        for k,v in outsidei.items():
         j=k+v
        resulti.append(j)
   pass

   def interceptdata(self):
        stri0=stri[a:b]
        pass

   def sendemail(self):
    
        pass
pass
























# list1=[]
# for name in ws.iter_rows(min_row=1, max_row=79, min_col=1, max_col=1, values_only=True): #转化为列表
#    list10=list(name)
#    str1=str(list10)
#    str10=str1
#    str11='您好，'
#    str12=str11+str10
#    list1.append(str12) 
# pass

# list2=[]
# for sex in ws.iter_rows(min_row=1, max_row=79, min_col=22, max_col=22, values_only=True):
#    str2=' '
#    list20=list(sex) 
#    if list20==['男']:
#       str2='先生。' 
#       pass
#    elif list20==['女']:
#       str2='女士。'
#       pass
#    else:
#       str2='性别异常，请检查'
#       pass
#    list2.append(str2)   
# pass


# outside1=dict(zip(list1,list2))
# result1=[]
# for k,v in outside1.items():
#     j=k+v
#     result1.append(j)
# pass


# list3=[]
# for jobname in ws.iter_rows(min_row=1, max_row=79, min_col=5, max_col=5, values_only=True):
#   list30=list(jobname)
#   str3=str(list30)
#   str30=str3[2:-2]
#   str31='您应聘的岗位是'
#   str32=str31+str30
#   list3.append(str32)
# pass 

# outside2=dict(zip(result1,list3))
# result2=[]
# for k,v in outside2.items():
#     j=k+v
#     result2.append(j)
# pass






# list4=[]
# for interviewtime in ws.iter_rows(min_row=1, max_row=79, min_col=24, max_col=24, values_only=True):
#    list40=list(interviewtime)
#    str4=str(list40)
#    str40=str4[2:-2]
#    str41='，您的面试时间是：'
#    str42=str41+str40
#    list4.append(str42)
# pass

# list5=[]
# for interviewid in ws.iter_rows(min_row=1, max_row=79, min_col=25, max_col=25, values_only=True):
#    list50=list(interviewid)
#    str5=str(list50)
#    str50=str5[1:-1]
#    str51=',您的面试会议号码是：'
#    str52=str51+str50
#    list5.append(str52)
# pass

# outside3=dict(zip(list4,list5))
# result3=[]
# for k,v in outside3.items():
#     j=k+v
#     result3.append(j)
# pass

# outside4=dict(zip(result2,result3))
# result4=[]
# for k,v in outside4.items():
#     j=k+v
#     result4.append(j)
# pass


# list6=[]
# for emailname in ws.iter_rows(min_row=1, max_row=79, min_col=21, max_col=21, values_only=True):
#    list60=list(emailname) 
#    str6=str(list60)
#    str60=str6[1:-1]
#    list6.append(str60)
# pass
# outside5=dict(zip(result4,list6))
# for k,v in outside5.items():
#  pass
 
# import yagmail
# mail_host='smtp.qq.com'
# mail_pass='vdudxzcckwzbcdaj'#授权码
# sender='1756336885@qq.com'
# subjects='面试安排'
# import yagmail
#  mail_host='smtp.163.com'
#  mail_pass='WCDDCXKFRTSPSNMF'#授权码
#  sender='niejianghai1314@163.com'
#  subjects='面试安排'
#  yag=yagmail.SMTP(user=sender,password=mail_pass,host=mail_host)
#  try:
#      yag.send(to=v,subject=subjects,contents=k)
#      print('发送成功')
#  except Exception as mag:
#      print('发送失败',mag)
# pass